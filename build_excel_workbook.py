"""
Builds the Excel deliverable: an "Orders" sheet with the full order-level
fact table (as a proper Excel Table), and a "Summary" sheet built entirely
with live formulas (SUMIFS/COUNTIFS/AVERAGEIFS) referencing the Orders sheet
— nothing here is a hardcoded Python-computed number.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(__file__)
fact = pd.read_csv(os.path.join(BASE, "data", "order_fact_table.csv"))

NAVY = "1F3864"
LIGHTBG = "F2F4F8"
ARIAL = "Arial"

wb = Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: Orders (raw fact table as a real Excel Table)
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Orders"

headers = list(fact.columns)
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)

for row in fact.itertuples(index=False):
    ws.append(list(row))

n_rows = len(fact) + 1
n_cols = len(headers)
last_col_letter = get_column_letter(n_cols)

for row in ws.iter_rows(min_row=2, max_row=n_rows, min_col=1, max_col=n_cols):
    for cell in row:
        cell.font = Font(name=ARIAL, size=10)

# number formats
col_idx = {name: i + 1 for i, name in enumerate(headers)}
for col_name, fmt in [
    ("price", '$#,##0.00'), ("freight_value", '$#,##0.00'), ("order_value", '$#,##0.00'),
]:
    letter = get_column_letter(col_idx[col_name])
    for cell in ws[letter][1:]:
        cell.number_format = fmt

table_ref = f"A1:{last_col_letter}{n_rows}"
tbl = Table(displayName="Orders", ref=table_ref)
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

for i, header in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(i)].width = max(12, len(header) + 2)

print(f"Orders sheet: {n_rows-1:,} rows written")

# ---------------------------------------------------------------------------
# Sheet 2: Summary (all formulas, no hardcoded results)
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False


def title(cell_ref, text, size=14):
    c = ws2[cell_ref]
    c.value = text
    c.font = Font(name=ARIAL, bold=True, size=size, color=NAVY)


def label(cell_ref, text, bold=False):
    c = ws2[cell_ref]
    c.value = text
    c.font = Font(name=ARIAL, bold=bold)


def formula(cell_ref, f, number_format=None):
    c = ws2[cell_ref]
    c.value = f
    c.font = Font(name=ARIAL)
    if number_format:
        c.number_format = number_format


col = col_idx
COL = {name: get_column_letter(i) for name, i in col_idx.items()}

title("A1", "E-Commerce Delivery Analytics — Summary")
ws2["A2"] = "All figures are live formulas referencing the Orders sheet."
ws2["A2"].font = Font(name=ARIAL, italic=True, size=9, color="666666")

# --- Overall KPIs ---
title("A4", "Overall KPIs", size=12)
label("A5", "Total delivered orders")
formula("B5", f"=COUNTA(Orders!{COL['order_id']}2:{COL['order_id']}100000)")

label("A6", "Late orders")
formula("B6", f"=SUM(Orders!{COL['is_late']}2:{COL['is_late']}100000)")

label("A7", "% of orders late")
formula("B7", "=B6/B5", "0.0%")

label("A8", "Avg. review score — on time")
formula("B8", f"=AVERAGEIFS(Orders!{COL['review_score']}2:{COL['review_score']}100000,"
              f"Orders!{COL['is_late']}2:{COL['is_late']}100000,0)", "0.00")

label("A9", "Avg. review score — late")
formula("B9", f"=AVERAGEIFS(Orders!{COL['review_score']}2:{COL['review_score']}100000,"
              f"Orders!{COL['is_late']}2:{COL['is_late']}100000,1)", "0.00")

label("A10", "Avg. order value")
formula("B10", f"=AVERAGE(Orders!{COL['order_value']}2:{COL['order_value']}100000)", '$#,##0.00')

label("A11", "Total order value, late deliveries")
formula("B11", f"=SUMIFS(Orders!{COL['order_value']}2:{COL['order_value']}100000,"
               f"Orders!{COL['is_late']}2:{COL['is_late']}100000,1)", '$#,##0.00')

# --- Late rate by seller state (top states by volume) ---
title("A13", "Late Delivery Rate by Seller State (top 10 by order volume)", size=12)
label("A14", "Seller State", bold=True)
label("B14", "Total Orders", bold=True)
label("C14", "Late Orders", bold=True)
label("D14", "Late Rate", bold=True)

top_states = fact["seller_state"].value_counts().head(10).index.tolist()
for i, state in enumerate(top_states):
    r = 15 + i
    ws2[f"A{r}"] = state
    ws2[f"A{r}"].font = Font(name=ARIAL)
    formula(f"B{r}", f'=COUNTIFS(Orders!{COL["seller_state"]}2:{COL["seller_state"]}100000,A{r})')
    formula(f"C{r}", f'=COUNTIFS(Orders!{COL["seller_state"]}2:{COL["seller_state"]}100000,A{r},'
                     f'Orders!{COL["is_late"]}2:{COL["is_late"]}100000,1)')
    formula(f"D{r}", f"=C{r}/B{r}", "0.0%")

state_end_row = 14 + len(top_states)

# --- Late rate by category (top 10 by volume) ---
cat_start = state_end_row + 2
title(f"A{cat_start}", "Late Delivery Rate by Product Category (top 10 by order volume)", size=12)
label(f"A{cat_start+1}", "Category", bold=True)
label(f"B{cat_start+1}", "Total Orders", bold=True)
label(f"C{cat_start+1}", "Late Orders", bold=True)
label(f"D{cat_start+1}", "Late Rate", bold=True)

top_cats = fact["category"].value_counts().head(10).index.tolist()
for i, cat in enumerate(top_cats):
    r = cat_start + 2 + i
    ws2[f"A{r}"] = cat
    ws2[f"A{r}"].font = Font(name=ARIAL)
    formula(f"B{r}", f'=COUNTIFS(Orders!{COL["category"]}2:{COL["category"]}100000,A{r})')
    formula(f"C{r}", f'=COUNTIFS(Orders!{COL["category"]}2:{COL["category"]}100000,A{r},'
                     f'Orders!{COL["is_late"]}2:{COL["is_late"]}100000,1)')
    formula(f"D{r}", f"=C{r}/B{r}", "0.0%")

cat_end_row = cat_start + 1 + len(top_cats)

for letter, width in [("A", 34), ("B", 14), ("C", 14), ("D", 12)]:
    ws2.column_dimensions[letter].width = width

# --- Chart: late rate by seller state ---
chart = BarChart()
chart.title = "Late Delivery Rate by Seller State"
chart.y_axis.title = "Late rate"
chart.x_axis.title = "Seller state"
data_ref = Reference(ws2, min_col=4, min_row=14, max_row=state_end_row)
cats_ref = Reference(ws2, min_col=1, min_row=15, max_row=state_end_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height, chart.width = 8, 16
ws2.add_chart(chart, f"F14")

out_path = os.path.join(BASE, "Ecommerce_Delivery_Analytics.xlsx")
wb.save(out_path)
print(f"Workbook saved: {out_path}")
